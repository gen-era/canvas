import openpyxl
from datetime import datetime
from canvas.models import Institution  # Adjust the import based on your app name


def generate_data_list(file_path):
    # Define corresponding headers for each field
    prot_id_list = ["Genetik Protokol No", "Protokol No"]
    inst_list = ["Kurum", "Bölge"]
    arrival_date_list = ["Kayıt Tarihi", "Test Eklenme Tarihi"]

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Read the header row (assuming the first row is the header)
    headers = [cell.value for cell in ws[1]]

    # Function to find the column index for a given list of possible headers
    def find_column(header_options):
        for option in header_options:
            if option in headers:
                return headers.index(option) + 1  # openpyxl is 1-indexed
        return None

    # Find column indices for each field
    prot_id_col = find_column(prot_id_list)
    inst_col = find_column(inst_list)
    arrival_date_col = find_column(arrival_date_list)

    # Check if all required columns are found
    required_columns = {
        "prot_id": prot_id_col,
        "inst": inst_col,
        "arrival_date": arrival_date_col,
    }

    missing_columns = [key for key, val in required_columns.items() if val is None]
    if missing_columns:
        raise ValueError(
            f"Missing columns in the Excel file: {', '.join(missing_columns)}"
        )

    # Fetch existing institution names from the database
    existing_institutions = set(Institution.objects.values_list("name", flat=True))

    data_list = []

    # Iterate over the rows starting from the second row
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Extract data based on column indices
        prot_id = row[prot_id_col - 1]
        inst = row[inst_col - 1]
        arrival = row[arrival_date_col - 1]

        if inst in existing_institutions:
            try:
                inst_validated = Institution.objects.filter(name__icontains=inst[:5])
            except:
                inst_validated = None
            if len(inst_validated) > 1:
                inst_validated = None
            elif inst_validated:
                inst_validated = inst_validated.first()
            else:
                inst_validated = None
        else:
            inst_validated = None

        # Format 'arrival_date' if it's a datetime object
        if isinstance(arrival, datetime):
            arrival_formatted = arrival.strftime("%Y-%m-%d")
        elif isinstance(arrival, str):
            # If it's already a string, you might want to validate or reformat it
            try:
                # Attempt to parse and reformat
                arrival_dt = datetime.strptime(arrival, "%Y-%m-%d")
                arrival_formatted = arrival_dt.strftime("%Y-%m-%d")
            except ValueError:
                # If parsing fails, keep it as is or handle accordingly
                arrival_formatted = arrival
        else:
            # Handle other possible types (e.g., None)
            arrival_formatted = ""

        # Create a dictionary for the current row
        row_dict = {
            "prot_id": prot_id,
            "inst": inst_validated,  # Use the validated institution name
            "arrival_date": arrival_formatted,  # Use the formatted date string
        }

        data_list.append(row_dict)

    return data_list
